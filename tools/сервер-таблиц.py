# -*- coding: utf-8 -*-
"""КРОШЕЧНЫЙ СЕРВЕР, ЧТОБЫ ОТДАТЬ ТАБЛИЦЫ КАРТ В БРАУЗЕР.

Зачем он нужен. Данные карт надо вставить в Google Таблицы. Гнать их через
переписку — это сотни килобайт текста, который никому не нужен глазами; гонять
их руками по ячейкам — часы. Проще отдать браузеру по локальному адресу: сама
страница Google Таблиц забирает лист, кладёт в буфер обмена и вставляет.

Почему http на localhost работает со https-страницы: Chrome считает
http://127.0.0.1 доверенным источником, и правило про смешанное содержимое на
него не распространяется. Заголовок Access-Control-Allow-Origin разрешает
запрос со стороны docs.google.com.

Данные читаются из той же книги, что и выгрузка (docs/КАРТЫ — все таблицы.xlsx):
один источник, два получателя — файл и браузер.

Запуск: python tools/сервер-таблиц.py [порт]
"""
import json
import sys
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

from openpyxl import load_workbook

КНИГА = "docs/КАРТЫ — все таблицы.xlsx"
ПОРТ = int(sys.argv[1]) if len(sys.argv) > 1 else 8765


def листы():
    """Книга целиком: имя листа → строка TSV со шапкой и всеми строками."""
    wb = load_workbook(КНИГА)
    out = {}
    for имя in wb.sheetnames:
        ws = wb[имя]
        строки = []
        for row in ws.iter_rows(values_only=True):
            ячейки = []
            for v in row:
                if v is None:
                    ячейки.append("")
                    continue
                s = str(v)
                # TAB И ПЕРЕВОД СТРОКИ ВНУТРИ ЯЧЕЙКИ ЛОМАЮТ ВСТАВКУ: таблица
                # разъедется по колонкам и строкам. Заменяем на пробел — текст
                # карт всё равно идёт одним абзацем.
                ячейки.append(s.replace("\t", " ").replace("\n", " ").replace("\r", " "))
            while ячейки and ячейки[-1] == "":
                ячейки.pop()
            строки.append("\t".join(ячейки))
        out[имя] = "\n".join(строки)
    return out


ДАННЫЕ = json.dumps(листы(), ensure_ascii=False)


class Обработчик(BaseHTTPRequestHandler):
    def do_GET(self):
        # ОДИН ЛИСТ ПРОСТЫМ ТЕКСТОМ: /?sheet=<имя>. Так страница localhost —
        # у неё, в отличие от docs.google.com, нет запрета на сторонние запросы —
        # кладёт лист в буфер обмена, а Google Таблицы его вставляют. Данные не
        # проходят ни через переписку, ни через файловый диалог.
        from urllib.parse import urlparse, parse_qs, unquote
        q = parse_qs(urlparse(self.path).query)
        имя = unquote(q.get("sheet", [""])[0])
        if имя:
            листы_все = json.loads(ДАННЫЕ)
            тело = листы_все.get(имя, "НЕТ ТАКОГО ЛИСТА: " + имя).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Content-Length", str(len(тело)))
            self.end_headers()
            self.wfile.write(тело)
            return
        тело = ДАННЫЕ.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(тело)))
        self.end_headers()
        self.wfile.write(тело)

    def log_message(self, *args):
        pass            # тишина: сервер живёт минуту и не должен сорить в лог


if __name__ == "__main__":
    имена = list(json.loads(ДАННЫЕ).keys())
    print("листов:", len(имена), "->", ", ".join(имена))
    print(f"слушаю http://127.0.0.1:{ПОРТ}/  (Ctrl+C чтобы остановить)")
    ThreadingHTTPServer(("127.0.0.1", ПОРТ), Обработчик).serve_forever()
